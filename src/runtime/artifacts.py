"""
Артефакты-для-отдачи: агент ПРОИЗВОДИТ файл (таблица → .xlsx/.csv) и ДОСТАВЛЯЕТ его
пользователю как скачиваемый файл, а не вываливает CSV-текст в чат.

Закрывает дыру: `write_file`/`stash_export_csv` либо пишут на диск сервера «в никуда», либо
возвращают CSV строкой → юзер должен копипастить. Здесь файл пишется в `artifacts/<id>/<имя>`
(стабильный, server-servable каталог; относительный → резолвится в workdir упакованного .app или
cwd в деве, как `data/`), а МЕТА регистрируется в run_context → сервер отдаёт `GET /artifact/{id}`,
GUI рисует кнопку «скачать», CLI печатает путь.

Тул НАТИВНЫЙ (не python_exec) → песочница (deny file-write вне /tmp) ему не мешает; пишет он
ТОЛЬКО в свой artifacts-каталог, имя файла санитизируется (анти-traversal).
"""
from __future__ import annotations

import csv
import io
import re
import uuid
from pathlib import Path

from langchain_core.tools import StructuredTool
from pydantic import BaseModel, Field

from src.runtime import run_context


def artifact_dir() -> Path:
    """Каталог артефактов (относительный → workdir/.app или cwd, как data/). Создаётся при записи."""
    d = Path("artifacts")
    d.mkdir(parents=True, exist_ok=True)
    return d


def _safe_name(name: str, default_ext: str = ".xlsx") -> str:
    """Имя файла без путей и мусора. Расширение по умолчанию .xlsx, если не задано осмысленно."""
    stem = Path(str(name or "export").strip()).name            # срезаем любые ../ и каталоги
    stem = re.sub(r"[^\w.\- ]", "_", stem).strip() or "export"  # только безопасные символы
    if not re.search(r"\.[A-Za-z0-9]{1,5}$", stem):
        stem += default_ext
    return stem


# Реестр по (run_id, имя) → id+путь: повторный экспорт ТОГО ЖЕ имени в одном прогоне добавляет
# ЛИСТ в ту же книгу (мультитабличный отчёт), а не плодит файлы. Чистится по выходе из прогона.
_books: dict[tuple, dict] = {}
run_context.register_cleanup(lambda rid: [_books.pop(k, None) for k in list(_books) if k[0] == rid])


def write_table(filename: str, columns: list, rows: list, sheet: str = "Sheet1") -> dict:
    """Записать таблицу в .xlsx (openpyxl) или .csv (fallback). Вернуть мета {id,name,path,nrows}.
    Повтор с тем же filename в этом прогоне → добавляет лист к существующей книге."""
    rid = run_context.current_run_id() or "_default"
    name = _safe_name(filename)
    columns = [str(c) for c in (columns or [])]
    rows = [[("" if c is None else c) for c in r] for r in (rows or [])]
    key = (rid, name.lower())

    existing = _books.get(key)
    aid = existing["id"] if existing else uuid.uuid4().hex
    adir = artifact_dir() / aid
    adir.mkdir(parents=True, exist_ok=True)
    path = adir / name

    try:
        import openpyxl
        if existing and path.suffix.lower() == ".xlsx" and path.exists():
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # убрать дефолтный пустой лист
        title = re.sub(r"[\\/?*\[\]:]", "_", str(sheet or "Sheet1"))[:31] or "Sheet1"
        base, i = title, 2
        while title in wb.sheetnames:  # уникальное имя листа
            title = f"{base}_{i}"[:31]; i += 1
        ws = wb.create_sheet(title)
        if columns:
            ws.append(columns)
        for r in rows:
            ws.append(list(r))
        if path.suffix.lower() != ".xlsx":     # имя без .xlsx, но движок есть → принудительно .xlsx
            path = path.with_suffix(".xlsx"); name = path.name
        wb.save(path)
        kind = "xlsx"
    except ImportError:
        # Fallback без openpyxl: .csv (Excel открывает штатно). Мультилист не поддержан — один файл.
        path = path.with_suffix(".csv"); name = path.name
        with open(path, "w", encoding="utf-8-sig", newline="") as f:  # utf-8-sig → кириллица в Excel
            w = csv.writer(f, delimiter=";")
            if columns:
                w.writerow(columns)
            w.writerows(rows)
        kind = "csv"

    meta = {"id": aid, "name": name, "path": str(path), "nrows": len(rows), "kind": kind}
    _books[key] = {"id": aid, "path": str(path)}
    run_context.artifact_emit(meta)
    return meta


def save_artifact_file(filename: str, content) -> dict:
    """Зарегистрировать ПРОИЗВОЛЬНЫЙ файл (сырой текст/байты) как доставляемый артефакт — чтобы любой
    файловый тул (write_file/stash_export_csv), а не только export_table, доходил до пользователя
    кнопкой/путём. Пишет в artifacts/<id>/<имя> (контейнер, не cwd) + регистрирует. Живой eval
    вскрыл: агент берёт write_file, пишет в cwd и говорит «доставить нельзя» — структурно убираем."""
    aid = uuid.uuid4().hex
    name = _safe_name(filename, default_ext=".txt")
    adir = artifact_dir() / aid
    adir.mkdir(parents=True, exist_ok=True)
    path = adir / name
    if isinstance(content, bytes):
        path.write_bytes(content)
    else:
        path.write_text(str(content), encoding="utf-8")
    meta = {"id": aid, "name": name, "path": str(path),
            "nrows": str(content).count("\n") + 1 if not isinstance(content, bytes) else 0, "kind": path.suffix.lstrip(".") or "file"}
    run_context.artifact_emit(meta)
    return meta


def resolve_artifact(artifact_id: str) -> Path | None:
    """id → путь к файлу для отдачи (GET /artifact/{id}). Жёсткая валидация id (анти-traversal):
    только hex, каталог artifacts/<id>/ должен существовать, отдаём единственный файл в нём."""
    if not re.fullmatch(r"[0-9a-f]{8,40}", str(artifact_id or "")):
        return None
    adir = artifact_dir() / artifact_id
    if not adir.is_dir():
        return None
    files = [p for p in adir.iterdir() if p.is_file()]
    return files[0] if files else None


class _Export(BaseModel):
    filename: str = Field(description="File name FOR THE USER, e.g. 'procurement_2025'. Extension "
                          "optional (defaults to .xlsx).")
    columns: list[str] = Field(description="Column headers (the first row).")
    rows: list[list[str]] = Field(description="Data rows; each row a list of cell values matching "
                                  "columns order. Numbers as plain strings ('2.97'), no thousands "
                                  "separators.")
    sheet: str = Field(default="Sheet1", description="Sheet/tab name. To put SEVERAL tables in one "
                       "workbook, call again with the SAME filename and a different sheet name.")


def make_export_tool() -> StructuredTool:
    def _export(filename: str, columns: list, rows: list, sheet: str = "Sheet1") -> str:
        try:
            m = write_table(filename, columns, rows, sheet)
        except Exception as e:  # noqa: BLE001
            return f"[export failed] {type(e).__name__}: {e}"
        return (f"Файл '{m['name']}' ({m['nrows']} строк, лист '{sheet}') готов и будет доставлен "
                f"пользователю как СКАЧИВАЕМЫЙ файл (в GUI — кнопка, в терминале — сохранённый путь). "
                f"НЕ вставляй содержимое таблицы как текст/CSV в ответ — только короткое описание, "
                f"что внутри файла.")

    return StructuredTool.from_function(
        func=_export, name="export_table", args_schema=_Export,
        description="Produce a DOWNLOADABLE spreadsheet file (.xlsx) and deliver it to the user. "
                    "Use this WHENEVER the user asks for a file / Excel / CSV / table export — the "
                    "user gets a real download button. Do NOT paste CSV/table text into the chat "
                    "telling them to copy-paste; call this instead. Call multiple times with the same "
                    "filename to add more sheets to one workbook.",
    )
